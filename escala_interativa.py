
import streamlit as st
import pandas as pd
import random
from collections import defaultdict
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

st.set_page_config(layout="centered")
st.title("Gerador de Escala Semanal - Modelo com Mesclas Ajustadas")

dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
turnos = ['Manhã', 'Tarde']
vies_turno = {'Jack': 'Tarde'}

data_inicio = st.date_input("Data de início da semana (segunda-feira):", value=datetime.today(), format="DD/MM/YYYY")
data_inicio_real = data_inicio - timedelta(days=data_inicio.weekday())
data_fim = data_inicio_real + timedelta(days=4)

usar_feriados = st.checkbox("Deseja excluir feriados nesta semana?", value=False)
feriados = []
if usar_feriados:
    feriados = st.date_input("Selecione os feriados:", value=[], format="DD/MM/YYYY", key="feriados", help="Datas selecionadas serão marcadas como feriado.", disabled=not usar_feriados)

datas_semana = []
for i in range(5):
    data = data_inicio_real + timedelta(days=i)
    dia_label = f"{dias_semana[i]} ({data.strftime('%d/%m')})"
    is_feriado = data in feriados
    datas_semana.append((dia_label, data, is_feriado))

pessoas_manha = st.number_input("Pessoas por turno da Manhã", min_value=2, max_value=10, value=4)
pessoas_tarde = st.number_input("Pessoas por turno da Tarde", min_value=2, max_value=10, value=2)
capacidade = {'Manhã': pessoas_manha, 'Tarde': pessoas_tarde}
num_pessoas = st.number_input("Quantos agentes?", min_value=2, max_value=20, value=6)

st.subheader("Agentes e seus turnos semanais")
nomes = []
limites = {}
restricao_dia = {}

with st.form("formulario_completo"):
    for i in range(int(num_pessoas)):
        col1, col2, col3 = st.columns([2, 1, 2])
        nome = col1.text_input("Agente", key=f"nome_{i}")
        turnos_semanais = col2.number_input("Turnos", min_value=1, max_value=50, value=1, key=f"turno_{i}")
        unica_vez = col3.checkbox("Máx. 1 turno por dia", value=False, key=f"restricao_{i}")
        if nome:
            nomes.append(nome)
            limites[nome] = turnos_semanais
            restricao_dia[nome] = unica_vez
    gerar = st.form_submit_button("Gerar Escala")

def gerar_escala():
    escala = {}
    contagem = defaultdict(int)
    dias_embaralhados = datas_semana.copy()
    random.shuffle(dias_embaralhados)
    for dia_label, data, is_feriado in dias_embaralhados:
        for turno in turnos:
            if is_feriado:
                escala[(dia_label, turno)] = ["Feriado"]
                continue
            escala[(dia_label, turno)] = []
            while len(escala[(dia_label, turno)]) < capacidade[turno]:
                candidatos = []
                for pessoa in nomes:
                    if contagem[pessoa] >= limites[pessoa]:
                        continue
                    if pessoa in escala[(dia_label, turno)]:
                        continue
                    if restricao_dia.get(pessoa, False):
                        if any(pessoa in escala.get((dia_label, t), []) for t in turnos):
                            continue
                    candidatos.append(pessoa)
                random.shuffle(candidatos)
                if not candidatos:
                    return None, None
                pesados = []
                for pessoa in candidatos:
                    preferencia = vies_turno.get(pessoa)
                    if preferencia == turno:
                        pesados.extend([pessoa]*5)
                    else:
                        pesados.append(pessoa)
                escolhido = random.choice(pesados)
                escala[(dia_label, turno)].append(escolhido)
                contagem[escolhido] += 1
    return escala, contagem

def validar(escala, contagem):
    if not escala:
        return False
    jack_turnos = [t for (d, t), ps in escala.items() if 'Jack' in ps]
    if len(jack_turnos) == 5:
        if jack_turnos.count('Tarde') < 2:
            return False
    for pessoa, total in contagem.items():
        if 2 <= total <= 5:
            turnos_pessoa = [t for (d, t), ps in escala.items() if pessoa in ps]
            if len(set(turnos_pessoa)) < 2:
                return False
        if total > 5:
            dias_pessoa = [d for (d, t), ps in escala.items() if pessoa in ps]
            if len(set(dias_pessoa)) < 4:
                return False
    return True

if gerar:
    if not nomes or len(nomes) != len(set(nomes)):
        st.warning("Preencha todos os nomes corretamente e sem repetições.")
    else:
        escala, contagem = None, None
        for _ in range(1000):
            temp_escala, temp_contagem = gerar_escala()
            if validar(temp_escala, temp_contagem):
                escala, contagem = temp_escala, temp_contagem
                break

        if escala:
            st.success("Escala gerada com sucesso!")

            data = []
            inclui_eixos = any(len(p) == 4 for p in escala.values() if isinstance(p, list))

            for (dia_label, _, _) in datas_semana:
                for turno in turnos:
                    agentes = escala.get((dia_label, turno), [])
                    linha = {
                        'Dia': dia_label,
                        'Turno': turno,
                        'Agentes': 'Feriado' if agentes == ["Feriado"] else ', '.join(agentes),
                        'Anhanguera': '',
                        'Dom Pedro': ''
                    }
                    if inclui_eixos and isinstance(agentes, list) and len(agentes) == 4:
                        temp = agentes[:]
                        random.shuffle(temp)
                        linha['Anhanguera'] = ', '.join(temp[:2])
                        linha['Dom Pedro'] = ', '.join(temp[2:])
                    data.append(linha)

            df = pd.DataFrame(data)

            if inclui_eixos:
                if "Anhanguera" not in df.columns:
                    df["Anhanguera"] = ""
                if "Dom Pedro" not in df.columns:
                    df["Dom Pedro"] = ""

            contagem_df = pd.DataFrame.from_dict(contagem, orient='index', columns=['Turnos']).reset_index()
            contagem_df.columns = ['Agente', 'Turnos']

            st.markdown("### 📋 Escala da Semana")
            st.dataframe(df)

            st.markdown("### 📊 Turnos por Agente")
            st.dataframe(contagem_df)

            # Exportação com mesclas ajustadas
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Escala de Campo"

            titulo = f"Escala Semanal ({data_inicio_real.strftime('%d/%m')} - {data_fim.strftime('%d/%m/%Y')})"
            ws.merge_cells("A1:E1")
            ws["A1"] = titulo
            ws["A1"].font = Font(bold=True)
            ws["A1"].alignment = Alignment(horizontal='center')

            headers = ["Dia", "Turno", "Agentes"]
            if inclui_eixos:
                headers += ["Anhanguera", "Dom Pedro"]
            ws.append(headers)

            linha_inicio = 3
            idx = linha_inicio
            ultimo_dia = ""
            for row in df.itertuples():
                atual_dia = getattr(row, "Dia")
                is_same = atual_dia == ultimo_dia
                if is_same:
                    ws.append(["", getattr(row, "Turno"), row['Agentes'], row['Anhanguera'], row['Dom Pedro']] if inclui_eixos else ["", getattr(row, "Turno"), row['Agentes']])
                else:
                    ws.append([atual_dia, getattr(row, "Turno"), row['Agentes'], row['Anhanguera'], row['Dom Pedro']] if inclui_eixos else [atual_dia, getattr(row, "Turno"), row['Agentes']])
                
                ultimo_dia = atual_dia
                idx += 1

            ws.append([])
            ws.append(['Agente', 'Turnos'])
            for row in contagem_df.itertuples(index=False):
                ws.append(list(row))

            wb.save(output)
            output.seek(0)

            nome_arquivo = f"Escala Semanal_({data_inicio_real.strftime('%d-%m')}_{data_fim.strftime('%d-%m-%Y')}).xlsx"
            st.download_button("📥 Baixar Escala Formatada", data=output, file_name=nome_arquivo)
        else:
            st.error("Não foi possível gerar uma escala válida com os parâmetros definidos.")
